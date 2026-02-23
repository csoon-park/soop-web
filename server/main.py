"""
크랙 미션 매니저 - Backend
soopchat Python 라이브러리로 SOOP 채팅 연결
SSE로 프론트엔드에 실시간 이벤트 전달
"""
import asyncio
import json
import time
import os
import io
import hashlib
import secrets
import sqlite3
from datetime import datetime
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, Query, Depends, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from soopchat import SoopChat, Balloon, Adballoon, ChatMessage, Subscription, Mission
from soopchat.api import ApiService


# ─── 인증 시스템 (SQLite) ───

DB_PATH = os.path.join(os.path.dirname(__file__), "auth.db")

def init_db():
    """DB 초기화 및 기본 비밀번호 설정"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS auth (
            id INTEGER PRIMARY KEY,
            password_hash TEXT NOT NULL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            created_at REAL NOT NULL
        )
    """)
    # 비밀번호가 없으면 기본값 설정
    c.execute("SELECT COUNT(*) FROM auth")
    if c.fetchone()[0] == 0:
        default_hash = hashlib.sha256("lee0421@!".encode()).hexdigest()
        c.execute("INSERT INTO auth (id, password_hash) VALUES (1, ?)", (default_hash,))
    conn.commit()
    conn.close()

def verify_password(password: str) -> bool:
    """비밀번호 검증"""
    pw_hash = hashlib.sha256(password.encode()).hexdigest()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT password_hash FROM auth WHERE id = 1")
    row = c.fetchone()
    conn.close()
    return row is not None and row[0] == pw_hash

def change_password(new_password: str):
    """비밀번호 변경"""
    new_hash = hashlib.sha256(new_password.encode()).hexdigest()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("UPDATE auth SET password_hash = ? WHERE id = 1", (new_hash,))
    conn.commit()
    conn.close()

def create_session() -> str:
    """세션 토큰 생성"""
    token = secrets.token_hex(32)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # 오래된 세션 정리 (24시간)
    c.execute("DELETE FROM sessions WHERE created_at < ?", (time.time() - 86400,))
    c.execute("INSERT INTO sessions (token, created_at) VALUES (?, ?)", (token, time.time()))
    conn.commit()
    conn.close()
    return token

def validate_session(token: str) -> bool:
    """세션 토큰 검증"""
    if not token:
        return False
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT created_at FROM sessions WHERE token = ?", (token,))
    row = c.fetchone()
    conn.close()
    if not row:
        return False
    # 24시간 만료
    return (time.time() - row[0]) < 86400

def delete_session(token: str):
    """세션 삭제"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM sessions WHERE token = ?", (token,))
    conn.commit()
    conn.close()

# DB 초기화
init_db()


async def require_auth(request: Request):
    """인증 미들웨어 - 쿠키 또는 헤더에서 토큰 확인"""
    token = request.cookies.get("session_token") or request.headers.get("X-Session-Token", "")
    if not validate_session(token):
        return None
    return token

async def auth_guard(request: Request):
    """인증 필수 - 실패 시 401"""
    token = await require_auth(request)
    if not token:
        raise HTTPException(status_code=401, detail="로그인이 필요합니다")
    return token


# ─── 글로벌 상태 ───
class AppState:
    def __init__(self):
        self.client: Optional[SoopChat] = None
        self.connected = False
        self.streamer_id = ""
        self.results: list[dict] = []          # 미션 결과 목록
        self.templates: list[dict] = []        # 미션 템플릿
        self.auto_threshold = 0                # 자동등록 임계값
        self.logs: list[dict] = []             # 실시간 로그 (최대 200개)
        self.sse_queues: list[asyncio.Queue] = []  # SSE 구독자
        self._task: Optional[asyncio.Task] = None
        self._should_reconnect = False         # 자동 재연결 플래그

    def add_log(self, msg: str, log_type: str = "info"):
        entry = {
            "time": datetime.now().strftime("%p %I:%M:%S"),
            "message": msg,
            "type": log_type,
        }
        self.logs.insert(0, entry)
        if len(self.logs) > 200:
            self.logs = self.logs[:200]
        self.broadcast({"event": "log", "data": entry})

    def broadcast(self, data: dict):
        dead = []
        for q in self.sse_queues:
            try:
                q.put_nowait(data)
            except asyncio.QueueFull:
                dead.append(q)
        for q in dead:
            self.sse_queues.remove(q)

    def get_stats(self):
        total = len(self.results)
        done = sum(1 for r in self.results if r.get("done"))
        pending_count = total - done
        return {"total": total, "in_progress": pending_count, "done": done}


state = AppState()


@asynccontextmanager
async def lifespan(app: FastAPI):
    yield
    # 종료 시 연결 해제
    if state.client:
        await state.client.disconnect()


app = FastAPI(title="크랙 미션 매니저", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── 인증 API ───

@app.post("/api/login")
async def login(request: Request):
    body = await request.json()
    password = body.get("password", "")
    if not verify_password(password):
        return JSONResponse({"ok": False, "error": "비밀번호가 틀렸습니다"}, 401)
    token = create_session()
    resp = JSONResponse({"ok": True})
    resp.set_cookie("session_token", token, httponly=True, max_age=86400, samesite="lax")
    return resp

@app.post("/api/logout")
async def logout(request: Request):
    token = request.cookies.get("session_token", "")
    if token:
        delete_session(token)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("session_token")
    return resp

@app.get("/api/auth-check")
async def auth_check(request: Request):
    token = await require_auth(request)
    return {"ok": token is not None}

@app.post("/api/change-password")
async def change_password_api(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    current = body.get("current_password", "")
    new_pw = body.get("new_password", "")
    if not verify_password(current):
        return JSONResponse({"ok": False, "error": "현재 비밀번호가 틀렸습니다"}, 400)
    if len(new_pw) < 4:
        return JSONResponse({"ok": False, "error": "비밀번호는 최소 4자 이상이어야 합니다"}, 400)
    change_password(new_pw)
    return {"ok": True}


# ─── SSE 스트림 ───

@app.get("/api/events")
async def sse_events(request: Request, _=Depends(auth_guard)):
    queue = asyncio.Queue(maxsize=100)
    state.sse_queues.append(queue)

    async def event_generator():
        try:
            # 초기 상태 전송
            yield f"data: {json.dumps({'event': 'status', 'data': {'connected': state.connected, 'streamer_id': state.streamer_id, 'stats': state.get_stats()}})}\n\n"
            yield f"data: {json.dumps({'event': 'templates', 'data': state.templates})}\n\n"
            yield f"data: {json.dumps({'event': 'results', 'data': state.results})}\n\n"

            while True:
                if await request.is_disconnected():
                    break
                try:
                    data = await asyncio.wait_for(queue.get(), timeout=30)
                    yield f"data: {json.dumps(data, default=str)}\n\n"
                except asyncio.TimeoutError:
                    yield f": keepalive\n\n"
        finally:
            if queue in state.sse_queues:
                state.sse_queues.remove(queue)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ─── 스트리머 검색 ───

@app.get("/api/search-streamer")
async def search_streamer(request: Request, streamer_id: str = Query(...), _=Depends(auth_guard)):
    try:
        api = ApiService()
        data = api.get_socket_data(streamer_id)
        return {"ok": True, "streamer_id": streamer_id, "live": True}
    except Exception as e:
        return {"ok": False, "error": str(e), "streamer_id": streamer_id}


# ─── 연결/해제 ───

@app.post("/api/connect")
async def connect_streamer(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    streamer_id = body.get("streamer_id", "").strip()

    if not streamer_id:
        return JSONResponse({"ok": False, "error": "스트리머 ID를 입력하세요"}, 400)

    # 기존 연결 해제
    state._should_reconnect = False  # 기존 재연결 루프 중지
    if state.client:
        await state.client.disconnect()
        state.connected = False
    await asyncio.sleep(0.5)  # 기존 태스크 정리 대기

    state.streamer_id = streamer_id

    # 새 클라이언트 생성
    client = SoopChat(streamer_id)

    def on_connect(connected):
        state.connected = connected
        if connected:
            state.add_log(f"{streamer_id} 연결됨", "success")
        else:
            state.add_log(f"{streamer_id} 연결 해제됨", "warn")
        state.broadcast({"event": "status", "data": {
            "connected": state.connected,
            "streamer_id": state.streamer_id,
            "stats": state.get_stats(),
        }})

    def on_join(success):
        if success:
            state.add_log("채팅방 입장 성공!", "success")
        else:
            state.add_log("채팅방 입장 실패", "error")

    # 최근 도네이션 유저 추적 (채팅 메시지 연결용 - 별풍/애드/미션 모두)
    recent_donation_users = {}  # {user_id: {"result_id": ..., "time": ...}}

    def on_balloon(b: Balloon):
        result_id = _handle_donation("balloon", b.user.id, b.user.name, b.count, "", "")
        if result_id:
            recent_donation_users[b.user.id] = {"result_id": result_id, "time": time.time()}

    def on_adballoon(ab: Adballoon):
        result_id = _handle_donation("adballoon", ab.user.id, ab.user.name, ab.count, "", "")
        if result_id:
            recent_donation_users[ab.user.id] = {"result_id": result_id, "time": time.time()}

    def on_subscription(sub: Subscription):
        state.add_log(f"구독: {sub.user.name} ({sub.count}개월)", "info")
        state.broadcast({"event": "subscription", "data": {
            "user_id": sub.user.id,
            "user_nickname": sub.user.name,
            "count": sub.count,
        }})

    def on_mission(m: Mission):
        result_id = _handle_donation("mission", m.user.id, m.user.name, m.count, m.title, "")
        if result_id:
            recent_donation_users[m.user.id] = {"result_id": result_id, "time": time.time()}

    def on_chat(msg: ChatMessage):
        # 도네이션(별풍/애드/미션) 보낸 유저의 채팅이면 → 해당 결과에 메시지 연결
        user_id = msg.user.id
        if user_id in recent_donation_users:
            info = recent_donation_users[user_id]
            # 5초 이내의 채팅만 연결
            if time.time() - info["time"] < 5:
                rid = info["result_id"]
                for r in state.results:
                    if r["id"] == rid and not r.get("message"):
                        r["message"] = msg.message
                        state.broadcast({"event": "result_update", "data": r})
                        state.add_log(f"💬 {msg.user.name}: {msg.message}", "info")
                        break
            del recent_donation_users[user_id]

    def on_error(err):
        state.add_log(f"오류: {err}", "error")
        print(f"[ERROR] {err}")

    def on_raw(raw_repr):
        # keepalive(svc=0000) 관련 메시지만 출력
        if "0000" in raw_repr[:20]:
            print(f"[KEEPALIVE] pong received")
        # 별풍선(svc=0018) raw 메시지 디버그
        elif "0018" in raw_repr[:20]:
            print(f"[BALLOON RAW] {raw_repr[:500]}")

    client.on_connect(on_connect)
    client.on_join_channel(on_join)
    client.on_balloon(on_balloon)
    client.on_adballoon(on_adballoon)
    client.on_subscription(on_subscription)
    client.on_mission(on_mission)
    client.on_chat_message(on_chat)
    client.on_error(on_error)
    client.on_raw_message(on_raw)

    state.client = client
    state._should_reconnect = True

    # 백그라운드에서 연결 (자동 재연결 포함)
    async def run_client():
        retry_count = 0
        max_retries = 50  # 최대 50회 재연결 시도
        while state._should_reconnect and retry_count < max_retries:
            try:
                # 매번 새 클라이언트 생성 (재연결 시)
                if retry_count > 0:
                    new_client = SoopChat(streamer_id)
                    new_client.on_connect(on_connect)
                    new_client.on_join_channel(on_join)
                    new_client.on_balloon(on_balloon)
                    new_client.on_adballoon(on_adballoon)
                    new_client.on_subscription(on_subscription)
                    new_client.on_mission(on_mission)
                    new_client.on_chat_message(on_chat)
                    new_client.on_error(on_error)
                    new_client.on_raw_message(on_raw)
                    state.client = new_client
                    state.add_log(f"재연결 시도 #{retry_count}...", "warn")

                await state.client.connect()

            except Exception as e:
                err_msg = str(e)
                print(f"[DISCONNECT] {err_msg}")

                # 방송 종료인 경우 재연결 중단
                if "방송 중이 아닙니다" in err_msg:
                    state.add_log(f"방송이 종료되었습니다", "warn")
                    state._should_reconnect = False
                    break

                state.connected = False
                state.broadcast({"event": "status", "data": {
                    "connected": False,
                    "streamer_id": state.streamer_id,
                    "stats": state.get_stats(),
                }})

                if not state._should_reconnect:
                    break

                retry_count += 1
                wait_sec = min(5 * retry_count, 30)  # 5초, 10초... 최대 30초
                state.add_log(f"연결 끊김 → {wait_sec}초 후 재연결 ({retry_count}/{max_retries})", "warn")
                await asyncio.sleep(wait_sec)

        if retry_count >= max_retries:
            state.add_log("최대 재연결 횟수 초과. 수동으로 재연결하세요.", "error")
        state.connected = False
        state.broadcast({"event": "status", "data": {
            "connected": False,
            "streamer_id": state.streamer_id,
            "stats": state.get_stats(),
        }})

    state._task = asyncio.create_task(run_client())
    state.add_log(f"{streamer_id} 연결 시도 중...", "info")

    return {"ok": True, "streamer_id": streamer_id}


def _handle_donation(dtype: str, user_id: str, user_name: str, count: int, title: str, message: str = ""):
    """별풍선/애드벌룬/미션 수신 처리. 매칭 시 result_id 반환."""
    type_labels = {"balloon": "별풍선", "adballoon": "애드벌룬", "mission": "대결미션"}

    log_msg = f"{type_labels.get(dtype, dtype)}: {user_name}({user_id}) {count}개"
    if title:
        log_msg += f" [{title}]"
    if message:
        log_msg += f" 💬{message}"

    state.add_log(log_msg, dtype)

    # 결과 추가
    result_id = len(state.results) + 1
    result = {
        "id": result_id,
        "type": dtype,
        "user_id": user_id,
        "user_nickname": user_name,
        "count": count,
        "title": title,
        "message": message,
        "memo": "",
        "done": False,
        "matched_template": "",
        "time": datetime.now().strftime("%p %I:%M:%S"),
        "timestamp": time.time(),
    }

    # 템플릿 매칭 (정확히 일치)
    matched = False
    for tmpl in state.templates:
        if not tmpl.get("active", True):
            continue
        tmpl_type = tmpl.get("type", "all")
        if tmpl_type != "all" and tmpl_type != dtype:
            continue
        if count == tmpl.get("count", 0):
            result["matched_template"] = tmpl.get("name", "")
            matched = True
            break

    # 자동등록 임계값 (이상)
    if state.auto_threshold > 0 and count >= state.auto_threshold:
        result["matched_template"] = result.get("matched_template") or "자동등록"
        matched = True

    # 매칭된 것만 결과에 저장 (미매칭은 로그에만 기록)
    if not matched:
        return None

    state.results.insert(0, result)
    state.broadcast({"event": "result", "data": result})
    state.broadcast({"event": "stats", "data": state.get_stats()})
    return result_id


@app.post("/api/disconnect")
async def disconnect_streamer(request: Request, _=Depends(auth_guard)):
    state._should_reconnect = False  # 자동 재연결 중지
    if state.client:
        await state.client.disconnect()
    state.connected = False
    state.add_log("연결 해제됨", "warn")
    state.broadcast({"event": "status", "data": {
        "connected": False,
        "streamer_id": state.streamer_id,
        "stats": state.get_stats(),
    }})
    return {"ok": True}


# ─── 템플릿 (미션 등록) ───

@app.get("/api/templates")
async def get_templates(request: Request, _=Depends(auth_guard)):
    return {"ok": True, "templates": state.templates}


@app.post("/api/templates")
async def add_template(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    tmpl = {
        "id": len(state.templates) + 1,
        "name": body.get("name", ""),
        "count": body.get("count", 0),
        "type": body.get("type", "all"),        # all, balloon, adballoon, mission
        "collect_message": body.get("collect_message", False),
        "active": True,
    }
    state.templates.append(tmpl)
    state.broadcast({"event": "templates", "data": state.templates})
    state.add_log(f"미션 등록: {tmpl['name']} ({tmpl['count']}개)", "success")
    return {"ok": True, "template": tmpl}


@app.post("/api/templates/update")
async def update_template(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    tmpl_id = body.get("id")
    for tmpl in state.templates:
        if tmpl["id"] == tmpl_id:
            tmpl.update({k: v for k, v in body.items() if k != "id"})
            break
    state.broadcast({"event": "templates", "data": state.templates})
    return {"ok": True}


@app.post("/api/templates/delete")
async def delete_template(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    tmpl_id = body.get("id")
    state.templates = [t for t in state.templates if t["id"] != tmpl_id]
    state.broadcast({"event": "templates", "data": state.templates})
    return {"ok": True}


# ─── 미션 결과 ───

@app.get("/api/results")
async def get_results(request: Request, _=Depends(auth_guard)):
    return {"ok": True, "results": state.results, "stats": state.get_stats()}


@app.post("/api/results/toggle")
async def toggle_result(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    rid = body.get("id")
    for r in state.results:
        if r["id"] == rid:
            r["done"] = not r["done"]
            state.broadcast({"event": "result_update", "data": r})
            break
    state.broadcast({"event": "stats", "data": state.get_stats()})
    return {"ok": True}


@app.post("/api/results/memo")
async def update_memo(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    rid = body.get("id")
    memo = body.get("memo", "")
    for r in state.results:
        if r["id"] == rid:
            r["memo"] = memo
            state.broadcast({"event": "result_update", "data": r})
            break
    return {"ok": True}


@app.post("/api/results/delete")
async def delete_result(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    rid = body.get("id")
    state.results = [r for r in state.results if r["id"] != rid]
    state.broadcast({"event": "results", "data": state.results})
    state.broadcast({"event": "stats", "data": state.get_stats()})
    return {"ok": True}


@app.post("/api/results/clear")
async def clear_results(request: Request, _=Depends(auth_guard)):
    state.results = []
    state.broadcast({"event": "results", "data": state.results})
    state.broadcast({"event": "stats", "data": state.get_stats()})
    state.add_log("결과 초기화됨", "warn")
    return {"ok": True}


# ─── 자동등록 설정 ───

@app.post("/api/config")
async def update_config(request: Request, _=Depends(auth_guard)):
    body = await request.json()
    if "auto_threshold" in body:
        state.auto_threshold = int(body["auto_threshold"])
        state.add_log(f"자동등록 임계값: {state.auto_threshold}개", "info")
    return {"ok": True}


# ─── 내보내기 ───

@app.get("/api/export-excel")
async def export_excel(request: Request, type_filter: str = "", template_filter: str = "", _=Depends(auth_guard)):
    export_results = state.results
    if type_filter:
        export_results = [r for r in export_results if r.get("type") == type_filter]
    if template_filter:
        export_results = [r for r in export_results if r.get("matched_template") == template_filter]

    headers_row = ["유저ID", "닉네임", "개수", "타입", "매칭 미션", "메시지", "메모", "완료", "시간"]

    def write_rows(ws, rows):
        ws.append(headers_row)
        for r in rows:
            ws.append([
                r["user_id"],
                r["user_nickname"],
                r["count"],
                r["type"],
                r.get("matched_template", ""),
                r.get("message", ""),
                r.get("memo", ""),
                "완료" if r.get("done") else "진행중",
                r.get("time", ""),
            ])

    wb = Workbook()

    # 전체 시트
    ws_all = wb.active
    ws_all.title = "전체"
    write_rows(ws_all, export_results)

    # 미션별 시트
    template_names = []
    for t in state.templates:
        template_names.append(t["name"])

    for tname in template_names:
        matched = [r for r in export_results if r.get("matched_template") == tname]
        if matched:
            safe_name = tname[:31]  # 엑셀 시트명 최대 31자
            ws_t = wb.create_sheet(title=safe_name)
            write_rows(ws_t, matched)

    # 미매칭 (자동등록 등)
    unmatched = [r for r in export_results if not r.get("matched_template")]
    if unmatched:
        ws_u = wb.create_sheet(title="미매칭")
        write_rows(ws_u, unmatched)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=mission_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"},
    )


@app.get("/api/copy-ids")
async def copy_ids(request: Request, type_filter: str = "", separator: str = ", ", _=Depends(auth_guard)):
    results = state.results
    if type_filter:
        results = [r for r in results if r.get("type") == type_filter]
    ids = [r["user_id"] for r in results]
    unique_ids = list(dict.fromkeys(ids))
    return {"ok": True, "ids": separator.join(unique_ids), "count": len(unique_ids)}


# ─── 정적 파일 (프론트엔드) ───
frontend_dist = os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if os.path.exists(frontend_dist):
    app.mount("/", StaticFiles(directory=frontend_dist, html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.environ.get("PORT", 8000)), reload=False)
